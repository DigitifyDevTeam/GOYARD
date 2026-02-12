import openpyxl

try:
    wb = openpyxl.load_workbook('Prix goyard IA.xlsx')
    print('Sheet names:', wb.sheetnames)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n=== Sheet: {sheet_name} ===')
        print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
        
        # Read first 20 rows
        print('First 20 rows:')
        for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
            print(f'Row {i+1}: {row}')
        
        # Check if there are more rows
        if ws.max_row > 20:
            print(f'... and {ws.max_row - 20} more rows')
            
except Exception as e:
    print(f'Error: {e}')
