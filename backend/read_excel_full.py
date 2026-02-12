import openpyxl

try:
    wb = openpyxl.load_workbook('Prix goyard IA.xlsx')
    print('Sheet names:', wb.sheetnames)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n=== Sheet: {sheet_name} ===')
        print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
        
        # Read all rows
        print('All rows:')
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if any(cell is not None for cell in row):  # Only print non-empty rows
                print(f'Row {i+1}: {row}')
            
except Exception as e:
    print(f'Error: {e}')
